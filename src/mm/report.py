from pydantic import BaseModel, Field
from src.mm.models.deployment import Deployment, DeploymentResponse
from src.mm.models.affectedProduct import AffectedProduct, AffectedProductResponse
from src.mm.models.vulnerability import Vulnerability, VulnerabilityResponse
from bs4 import BeautifulSoup as bs
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import aiohttp
import asyncio
import calendar
import re


async def create_xl(rep: 'MonthlyReport'):
    wb = Workbook()
    sh = wb.create_sheet(title='KB Articles')
    sh.cell(1, 1, value="KB")
    sh.cell(1, 2, value="Catalog")
    sh.cell(1, 3, value="Release Date")
    sh.cell(1, 4, value="Title")
    sh.cell(1, 5, value="Summary")
    sh.cell(1, 6, value="Notes")
    sh.cell(1, 7, value="Known Issues")
    sh.cell(1, 8, value="Approval: Brandon")
    sh.cell(1, 9, value="Approval: Andrew")
    for i, kb in enumerate(rep.kbs, 2):
        sh.cell(i, 1, value=f'=HYPERLINK("{kb.url}", "{kb.kb}")')
        sh.cell(i, 2, value=f'=HYPERLINK("{kb.catalog}", "Catalog")')
        sh.cell(i, 3, value=f'{kb.releaseDate}')

        # prods = "\n".join(kb.unique_products())
        # sh.cell(i, 3, value=f'{kb.title}\nRelease Date: {kb.releaseDate[:10]}\nMax Severity: {kb.highest_severity()}\n\nProducts:\n{prods}')
        sh.cell(i, 4, value=f'{kb.title}')
        if kb.description != "":
            sh.cell(i, 5, value=f'{kb.description}')
        if len(kb.unique_super()) > 0:
            sh.cell(i, 6, value=f'\n\nSuperseded By:\n{kb.unique_super()}')

    sh2 = wb.create_sheet(title='Microsoft CVE')
    cve_list = []
    for x in rep.vulnerabilities:
        cve_entry = {'Latest': x.latestRevisionDate[:10],
                     'CVE': x.cveNumber,
                     'Title': x.cveTitle,
                     'CVSS Score': x.temporalScore,
                     'Attack Vector': x.get_av(),
                     'User Interaction': x.get_ui(),
                     'Impact': x.impact,
                     'Highest Severity': x.severity,
                     'Public': x.publiclyDisclosed,
                     'Assessment': x.latestSoftwareRelease,
                     'Exploited': x.exploited
                     }
        cve_list.append(cve_entry)
    df = pd.DataFrame(cve_list)
    df.sort_values(by='CVSS Score', ascending=False, inplace=True, ignore_index=True)
    for r in dataframe_to_rows(df, index=False, header=True):
        sh2.append(r)
    wb.save(filename=f"{rep.name}_{rep.year}-{rep.month:02d}.xlsx")


def unpack_office_kbs(rep: 'MonthlyReport'):
    doc = bs(rep.office_html, 'html.parser')
    tr = doc.find_all('tr')
    for r in tr:
        if r.find('td'):
            td = [x.get_text() for x in r.find_all('td')]
            product = td[0].strip()
            num = td[1].strip()
            num = re.search(r"\(KB(.*)\)", num)
            if num:
                num = num.group(1)
                if num in rep.unique_kb:
                    continue
                else:
                    new_kb = Kb(kb=num)
                    new_kb.releaseDate = get_second_tuesday_date(rep.year, rep.month)
                    new_kb.url = f'https://support.microsoft.com/help/{num}'
                    new_kb.products.append(product)
                    rep.kbs.append(new_kb)


def unpack_misc_kbs(rep: 'MonthlyReport'):
    doc = bs(rep.misc_html, 'html.parser')
    sections = doc.find('article').find_all('section', class_='ocpSection')
    for section in sections:
        if section.h2:
            if section.h2.get_text() == "More Information":
                month = section.section.h3.get_text()
                if month == rep.patch_day:
                    ul = section.section.find_all('ul', recursive=False)
                    for kb in ul:
                        kbtitle = re.search(r"\d*-\d* ([\S\s]*)[(]KB(\d*)[)]", kb.li.p.b.string)
                        if kbtitle:
                            num = kbtitle.group(2)
                        if num in rep.unique_kb:
                            continue
                        else:
                            new_kb = Kb(kb=num)
                            new_kb.releaseDate = get_second_tuesday_date(rep.year, rep.month)
                            new_kb.url = f'https://support.microsoft.com/help/{num}'
                            rep.kbs.append(new_kb)


def unpack_data(rep: 'MonthlyReport'):
    for d in rep.deployments:
        if re.search(r'.* .*', d.articleName):
            continue
        elif d.articleName not in rep.unique_kb:
            kb = Kb(kb=d.articleName, url=d.articleUrl, releaseDate=d.releaseDate[:10])
            kb.severity.append(d.severity)
            rep.kbs.append(kb)
            rep.unique_kb.append(d.articleName)
        else:
            kb = [x for x in rep.kbs if x.kb == d.articleName][0]
            kb.severity.append(d.severity)
    unpack_misc_kbs(rep)
    unpack_office_kbs(rep)


def get_specific_deployment_by_article(articleName: str):
    return f"https://api.msrc.microsoft.com/sug/v2.0/en-US/deployment/?%24orderBy=product+desc&%24filter=articleName+eq+%27{articleName}%27"


def get_specific_ap_by_cve(cveNumber: str):
    return f"https://api.msrc.microsoft.com/sug/v2.0/en-US/affectedProduct?%24filter=cveNumber+eq+%27{cveNumber}%27"


def get_specific_ap_by_id(ap_id: str):
    return f"https://api.msrc.microsoft.com/sug/v2.0/en-US/affectedProduct/{ap_id}"


def get_specific_vuln_by_cve(cveNumber: str):
    return f"https://api.msrc.microsoft.com/sug/v2.0/en-US/vulnerability?%24orderBy=cveNumber+desc&%24filter=cveNumber+eq+%27{cveNumber}%27"


def get_misc_url() -> str:
    return "https://support.microsoft.com/help/894199"


def get_catalog_url(articleName: str) -> str:
    # include "KB" letters in the search
    return f"https://www.catalog.update.microsoft.com/Search.aspx?q=KB{articleName}"


def get_catalog_inline_url(update_id: str) -> str:
    # "https://www.catalog.update.microsoft.com/ScopedViewInline.aspx?updateid=" + id + "#PackageDetails"
    return f"https://www.catalog.update.microsoft.com/ScopedViewInline.aspx?updateid={update_id}"


async def gather_deployment(session: aiohttp.ClientSession, rep: 'MonthlyReport'):
    url = rep.get_deployment_api_url(skip=0)
    async with session.get(url) as response:
        json = await response.json()
        resp = DeploymentResponse(**json)
        resp_list = resp.value
        if int(resp.count) > len(resp_list):
            pages = [x*len(resp.value) for x in [*range(1, (int(resp.count)//len(resp.value) + 1))]]
            for page in pages:
                async with session.get(rep.get_deployment_api_url(skip=page)) as p:
                    json = await p.json()
                    resp_list.extend(DeploymentResponse(**json).value)
        rep.deployments = resp_list


async def gather_ap(session: aiohttp.ClientSession, rep: 'MonthlyReport'):
    url = rep.get_affectedProduct_api_url(skip=0)
    async with session.get(url) as response:
        json = await response.json()
        resp = AffectedProductResponse(**json)
        resp_list = resp.value
        if int(resp.count) > len(resp_list):
            pages = [x*len(resp.value) for x in [*range(1, (int(resp.count)//len(resp.value) + 1))]]
            for page in pages:
                async with session.get(rep.get_affectedProduct_api_url(skip=page)) as p:
                    json = await p.json()
                    resp_list.extend(AffectedProductResponse(**json).value)
        rep.aps = resp_list


async def gather_vulnerability(session: aiohttp.ClientSession, rep: 'MonthlyReport'):
    url = rep.get_vulnerability_api_url(skip=0)
    async with session.get(url) as response:
        json = await response.json()
        resp = VulnerabilityResponse(**json)
        resp_list = resp.value
        if int(resp.count) > len(resp_list):
            pages = [x*len(resp.value) for x in [*range(1, (int(resp.count)//len(resp.value) + 1))]]
            for page in pages:
                async with session.get(rep.get_vulnerability_api_url(skip=page)) as p:
                    json = await p.json()
                    resp_list.extend(VulnerabilityResponse(**json).value)
        rep.vulnerabilities = resp_list


async def gather_misc(session: aiohttp.ClientSession, rep: 'MonthlyReport'):
    url = get_misc_url()
    async with session.get(url) as response:
        rep.misc_html = await response.text()


async def gather_office(session: aiohttp.ClientSession, rep: 'MonthlyReport'):
    url = rep.get_office_url()
    async with session.get(url) as response:
        rep.office_html = await response.text()


async def gather_inline(idn: str, session: aiohttp.ClientSession) -> dict:
    url = get_catalog_inline_url(idn)
    id_dict = {}
    async with session.get(url) as inline:
        html = await inline.text()
        doc = bs(html, 'html.parser')
        desc = doc.find('span', id='ScopedViewHandler_desc')
        if desc:
            id_dict['desc'] = desc.get_text(strip=True)
        prod = doc.find('div', id='productsDiv')
        if prod:
            prod = prod.get_text(strip=True, separator=";;")
            find = re.search(r'.*;;([\S\s]*)', prod)
            if find:
                prod = find.group(1).strip()
                prod = re.sub(r'\n', "", prod)
                prod = re.sub(r'\r', "", prod)
                prod = re.sub(r' +', " ", prod)
                id_dict['prod'] = prod
        id_dict['super'] = []
        div = doc.find('div', id='supersededbyInfo')
        if div:
            a = div.find_all('a')
            if a:
                for x in a:
                    link = x.get_text()
                    find = re.search(r'.*\((KB.*)\)', link)
                    if find:
                        id_dict['super'].append(find.group(1))
        id_dict['severity'] = []
        sev = doc.find('span', id='ScopedViewHandler_msrcSeverity')
        if sev:
            severity = sev.get_text(strip=True)
            id_dict['severity'].append(severity)
    return id_dict


async def gather_catalog(kb: 'Kb', session: aiohttp.ClientSession):
    url = get_catalog_url(kb.kb)
    async with session.get(url) as catalog:
        html = await catalog.text()
        doc = bs(html, 'html.parser')
        no_results = doc.find('span', id='ctl00_catalogBody_noResultText')
        if not no_results:
            kb.catalog = url
            table = doc.find('table', id='ctl00_catalogBody_updateMatches')
            rows = table.find_all('tr')
            ids = []
            for row in rows:
                if len(row['id']) > 10:
                    ids.append(asyncio.create_task(gather_inline(row['id'][:36], session)))
            results = await asyncio.gather(*ids)
            kb.description = results[0].get('desc')
            for result in results:
                if result.get('super'):
                    kb.superseded.extend(result.get('super'))
                if result.get('prod'):
                    kb.products.append(result.get('prod'))
                if result.get('severity'):
                    kb.severity.extend(result.get('severity'))


async def gather_catalogs(rep: 'MonthlyReport'):
    catalog = []
    async with aiohttp.ClientSession() as session:
        for kb in rep.kbs:
            catalog.append(asyncio.create_task(gather_catalog(kb, session)))
        await asyncio.gather(*catalog)


async def gather_title(kb: 'Kb', session: aiohttp.ClientSession):
    async with session.get(kb.url) as article:
        print(kb.kb)
        html = await article.text()
        doc = bs(html, 'html.parser')
        title = doc.find('title')
        if title:
            short_title = re.search(r'(.*) - .*$', title.text)
            if short_title:
                new_title = short_title.group(1)
            else:
                new_title = title.text
            kb.title = new_title


async def gather_titles(rep: 'MonthlyReport'):
    titles = []
    connector = aiohttp.TCPConnector(force_close=True, limit=150)
    async with aiohttp.ClientSession(connector=connector) as session:
        for kb in rep.kbs:
            titles.append(asyncio.create_task(gather_title(kb, session)))
        await asyncio.gather(*titles)


async def gather_data(report: 'MonthlyReport'):
    async with aiohttp.ClientSession() as session:
        tasks = [asyncio.create_task(gather_deployment(session, report)),
                 asyncio.create_task(gather_ap(session, report)),
                 asyncio.create_task(gather_vulnerability(session, report)),
                 asyncio.create_task(gather_misc(session, report)),
                 asyncio.create_task(gather_office(session, report))]
        await asyncio.gather(*tasks)



def get_second_tuesday_string(y: int, m: int) -> str:
    """Get a string representation for the chosen second Tuesday

    :param y: year
    :param m: month
    :return: string in form -> 'Tuesday, Month dd, yyyy'
    """
    c = calendar.Calendar()
    # d[3] == 1 means the 4th value in the d tuple should match 1, which is Tuesday
    # d[1] == month means the second value in the d tuple should match month
    # ending [1] means the second value in the returned list, which is the second Tuesday date tuple
    second = list(filter(lambda d: d[3] == 1 and d[1] == m, c.itermonthdays4(y, m)))[1]
    months = {1: "January", 2: "February", 3: "March", 4: "April", 5: "May", 6: "June", 7: "July", 8: "August",
              9: "September", 10: "October", 11: "November", 12: "December"}
    return f"Tuesday, {months[m]} {second[2]}, {y}"


def get_second_tuesday_date(y: int, m: int, delta: int = 0) -> str:
    """Get a string representation for the chosen second Tuesday

    :param y: year
    :param m: month
    :param delta: days before(-) or after(+) the second tuesday
    :return: string in form -> 'yyyy-mm-dd'
    """
    c = calendar.Calendar()
    # d[3] == 1 means the 4th value in the d tuple should match 1, which is Tuesday
    # d[1] == month means the second value in the d tuple should match month
    # ending [1] means the second value in the returned list, which is the second Tuesday date tuple
    second = list(filter(lambda d: d[3] == 1 and d[1] == m, c.itermonthdays4(y, m)))[1]
    return f"{second[0]}-{second[1]:02d}-{second[2]+delta:02d}"


class Kb(BaseModel):
    kb: str
    url: str = ""
    title: str = ""
    releaseDate: str = ""
    products: list[str] = Field(default_factory=list)
    severity: list[str] = Field(default_factory=list)
    description: str = ""
    catalog: str = ""
    superseded: list[str] = Field(default_factory=list)

    def highest_severity(self) -> str:
        if "Critical" in self.severity:
            return "Critical"
        elif "Important" in self.severity:
            return "Important"
        elif "Moderate" in self.severity:
            return "Moderate"
        else:
            return "N/A"

    def unique_products(self) -> list:
        return list(set(self.products))

    def unique_super(self) -> list:
        return list(set(self.superseded))


class MonthlyReport(BaseModel):
    """
    name: str
    year: int
    month: int
    """
    name: str
    year: int
    month: int
    misc_html: str = ""
    office_html: str = ""
    deployments: list[Deployment] = Field(default_factory=list)
    aps: list[AffectedProduct] = Field(default_factory=list)
    vulnerabilities: list[Vulnerability] = Field(default_factory=list)
    kbs: list[Kb] = Field(default_factory=list)
    unique_kb: list[str] = Field(default_factory=list)
    msrc_updates: int = Field(default=0)
    misc_updates: int = Field(default=0)
    office_updates: int = Field(default=0)
    msrc_cve: int = Field(default=0)

    @property
    def patch_day(self) -> str:
        return get_second_tuesday_string(self.year, self.month)

    @property
    def start(self):
        if self.month == 1:
            month = 12
            year = self.year - 1
            return f"{get_second_tuesday_date(year, month, 1)}"
        else:
            month = self.month
            year = self.year
            return f"{get_second_tuesday_date(year, month - 1, 1)}"

    @property
    def start_encoded(self):
        return f"{self.start}T00%3A00%3A00-06%3A00"

    @property
    def end(self):
        return f"{get_second_tuesday_date(self.year, self.month, 2)}"

    @property
    def end_encoded(self):
        return f"{self.end}T23%3A59%3A59-06%3A00"

    def get_vulnerability_api_url(self, skip: int) -> str:
        return f"https://api.msrc.microsoft.com/sug/v2.0/en-US/vulnerability?%24orderBy=cveNumber+asc&%24filter=%28releaseDate+gt+{self.start}T00%3A00%3A00-05%3A00+or+latestRevisionDate+gt+{self.start}T00%3A00%3A00-05%3A00%29+and+%28releaseDate+lt+{self.end}T23%3A59%3A59-05%3A00+or+latestRevisionDate+lt+{self.end}T23%3A59%3A59-05%3A00%29&$skip={str(skip)}"

    def get_affectedProduct_api_url(self, skip: int) -> str:
        return f"https://api.msrc.microsoft.com/sug/v2.0/en-US/affectedProduct?%24orderBy=releaseDate+desc&%24filter=%28releaseDate+gt+{self.start}T00%3A00%3A00-05%3A00%29+and+%28releaseDate+lt+{self.end}T23%3A59%3A59-05%3A00%29&$skip={str(skip)}"

    def get_deployment_api_url(self, skip: int) -> str:
        return f"https://api.msrc.microsoft.com/sug/v2.0/en-US/deployment/?%24orderBy=product+desc&%24filter=%28releaseDate+gt+{self.start}T00%3A00%3A00-06%3A00%29+and+%28releaseDate+lt+{self.end}T23%3A59%3A59-06%3A00%29&$skip={str(skip)}"

    def get_office_url(self) -> str:
        # "https://docs.microsoft.com/en-us/officeupdates/office-updates-msi"
        y = int(self.end[0:4])
        m = int(self.end[5:7])
        kb_dict = {
            1: "5002084",
            2: "5002085",
            3: "5002086",
            4: "5002087",
            5: "5002088",
            6: "5002089",
            7: "5002090",
            8: "5002091",
            9: "5002092",
            10: "5002093",
            11: "5002094",
            12: "5002095",
        }
        base_url = "https://support.microsoft.com/help/"
        if y == 2023:
            if 13 > m > 0:
                return f'{base_url}{kb_dict[m]}'
        else:
            print('Not the right year.')
            return ""

    async def run(self):
        print(f"Starting: {self.name}")
        print(f"Patch Tuesday: {self.patch_day}")
        print("Report Range:")
        print(self.start, "to", self.end)
        await gather_data(self)
        unpack_data(self)
        await gather_titles(self)
        await gather_catalogs(self)
        await create_xl(self)
